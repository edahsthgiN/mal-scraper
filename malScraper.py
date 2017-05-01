from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import urllib.request
import xml.etree.ElementTree as ET
 
username = input("Enter username of profile to be scraped (case sensitive, type '@exit' to exit): ")
 
while True:
	if username == "@exit":
		break
	username = username.rstrip()
	username = username.lstrip()
	url = "https://myanimelist.net/malappinfo.php?u=" + username + "&status=all&type=anime"
	userprofile = urllib.request.urlopen(url)

	tree = ET.parse(userprofile)
	root = tree.getroot()

	if not os.path.exists("User Scores"):
		os.makedirs("User Scores")
	try:
		os.remove("User Scores\\" + username + ".xlsx")
	except FileNotFoundError:		
		pass
		
	wb = Workbook()	
	ws = wb.active
	ws.title = username
		
		
	ws['A1'] = "Title"
	ws['B1'] = "User Score"
	ws['C1'] = "Status"
		
	row = 2
		
	for anime in root.findall('anime'):

		title = anime.find('series_title').text
		score = anime.find('my_score').text
		status = anime.find('my_status').text
			
		string_row = str(row)
			
		title_cell = 'A' + string_row
		score_cell = 'B' + string_row
		status_cell = 'C' + string_row
				
		if status == "1":
			ws[title_cell] = title
			ws[score_cell] = score
			ws[status_cell] = "Watching"
			row += 1	
		elif status == "2":
			ws[title_cell] = title
			ws[score_cell] = score
			ws[status_cell] = "Finished"
			row += 1
		elif status == "3":
			ws[title_cell] = title
			ws[score_cell] = score
			ws[status_cell] = "On Hold"			
			row += 1
		elif status == "4":
			ws[title_cell] = title
			ws[score_cell] = score
			ws[status_cell] = "Dropped"
			row += 1
	wb.save("User Scores\\" + username + ".xlsx")
	username = input("Enter username of profile: ")